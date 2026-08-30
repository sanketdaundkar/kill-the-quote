import json, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(1)

with open('/home/claude/aerchain-rfx/data/rfx/rfx_spec.json') as f:
    rfx = json.load(f)
with open('/home/claude/aerchain-rfx/data/rfx/base_prices.json') as f:
    base = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Quote"

ws.merge_cells('A1:F1')
ws['A1'] = "TechSource Distributors Pvt. Ltd. — Commercial Quotation"
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = "GSTIN: 27ABCDE1234F1Z5   |   Ref: TSD/2026/0912   |   Date: 28-Aug-2026"
ws['A3'] = "Valid till: 15-Sep-2026. All prices in INR, exclusive of GST (18% extra)."
ws['A4'] = "Delivery: 5-6 weeks ex-works Mumbai. Freight to Pune extra at actuals."

headers = ["S.No", "Product", "Brand/Model (indicative)", "MOQ", "Rate (INR, excl. GST)", "Remarks"]
hr = 6
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=hr, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5597")

# TechSource re-orders and re-describes items in their own words, doesn't use item codes.
# They quote ALL 30 lines but with variable +/-8% pricing and their own descriptions.
vendor_desc_map = {
    "IT-001": "Laptop - Core i5, 16GB/512GB SSD, 14\"",
    "IT-002": "Laptop - Core i3, 8GB/256GB SSD, 14\"",
    "IT-003": "Monitor 24\" FHD IPS",
    "IT-004": "Monitor 27\" QHD IPS (USB-C)",
    "IT-005": "Docking Station USB-C Dual 4K 90W",
    "IT-006": "Combo - Wireless Kb+Mouse",
    "IT-007": "Keyboard Wired USB",
    "IT-008": "Mouse USB Optical",
    "IT-009": "Webcam 1080p USB",
    "IT-010": "Headset USB w/ Mic (Noise Cancel)",
    "IT-011": "Headset Bluetooth",
    "IT-012": "Sleeve for 14-15\" Laptop",
    "IT-013": "Cable HDMI (USB-C) 2m",
    "IT-014": "Cable Cat6 3m",
    "IT-015": "Switch Managed 24-Port GbE",
    "IT-016": "Switch Managed 48-Port GbE",
    "IT-017": "AP WiFi6 Enterprise",
    "IT-018": "Rack 12U with PDU",
    "IT-019": "UPS Line-Interactive 1kVA",
    "IT-020": "UPS Online 3kVA",
    "IT-021": "SSD External 1TB USB-C",
    "IT-022": "HDD Portable 2TB USB3",
    "IT-023": "Hub USB-C 7-in-1",
    "IT-024": "Lock Cable for Laptop",
    "IT-025": "Stand Laptop Adjustable",
    "IT-026": "Arm Monitor Single Gas Spring",
    "IT-027": "Speakerphone Conf Room",
    "IT-028": "Camera PTZ Conf",
    "IT-029": "Scanner Barcode USB",
    "IT-030": "Printer Label Thermal",
}
order = list(vendor_desc_map.keys())
random.shuffle(order)

row = hr + 1
for sno, code in enumerate(order, start=1):
    price = round(base[code] * random.uniform(0.94, 1.08), -1)
    remark = ""
    if code == "IT-005":
        remark = "Subject to min order 100 units"
    ws.cell(row=row, column=1, value=sno)
    ws.cell(row=row, column=2, value=vendor_desc_map[code])
    ws.cell(row=row, column=3, value="Dell/HP/Lenovo equiv. - brand TBD on PO")
    ws.cell(row=row, column=4, value=1)
    ws.cell(row=row, column=5, value=price)
    ws.cell(row=row, column=6, value=remark)
    row += 1

ws.cell(row=row+1, column=1, value="Note: Warranty 3yr onsite on compute items only; peripherals carry standard 1yr OEM warranty.")
ws.cell(row=row+2, column=1, value="Payment: 50% advance, 50% on delivery (Net 45 negotiable for orders above INR 50L).")

for col, w in zip("ABCDEF", [6, 34, 30, 8, 18, 30]):
    ws.column_dimensions[col].width = w

ws2 = wb.create_sheet("Vendor Info")
ws2['A1'] = "Vendor Questionnaire Responses"
ws2['A1'].font = Font(bold=True, size=12)
qa_rows = [
    ("Q1", "OEM-authorized reseller for all quoted brands?", "Yes - Dell, HP, Lenovo, Cisco, APC"),
    ("Q2", "Single Pune drop-ship, one consolidated invoice?", "Yes"),
    ("Q3", "On-site 3-year warranty support in Pune?", "Yes"),
    ("Q4", "ISO 9001 / ISO 27001 certified?", "ISO 9001 only"),
    ("Q5", "Commit to 6-week delivery window for full order?", "Yes, subject to stock availability at PO time"),
]
ws2['A3'] = "Q ID"; ws2['B3'] = "Question"; ws2['C3'] = "Answer"
for r, (qid, q, a) in enumerate(qa_rows, start=4):
    ws2.cell(row=r, column=1, value=qid)
    ws2.cell(row=r, column=2, value=q)
    ws2.cell(row=r, column=3, value=a)
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 45

wb.save('/home/claude/aerchain-rfx/data/vendor_responses/vendor_a_techsource.xlsx')
print("saved, rows:", row)
